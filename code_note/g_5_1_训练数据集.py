# -*-coding:utf-8-*-

"""
（1）CSV文件
（2）JSON文件
（3）XLSX文件
（4）MySQL文件
"""

"""
CSV（Comma-Separated-Values），通常被称作逗号分隔值。CSV文件由任意数目的记录（行）组成，每条记录由一些字段（列）组
成，字段之间通常以逗号分隔，当然也可以用制表符等其他字符分割，所以CSV又被称为字符分割值。
Pandas 可以 直 接读 取CSV 文 件为 Series和DataFrame

"""

import pandas as pd


# 手动实现csv格式
def saveToCSV(rows, cols, data):
    f = open('mycsv.csv', 'w')
    f.write('A,B\n')  # 列名
    for i in range(rows):  # 对于每一行
        for j in range(cols):  # 对于每一列
            f.write(str(data.iloc[i][j]))  # 写入的数据应为字符型
            if j != cols - 1:  # 每行最后不写入逗号
                f.write(',')  # 设置分隔符
        f.write('\n')
    f.close()


# data = {'A': [1, 2, 3], 'B': [4, 5, 6]}
# df = pd.DataFrame(data)
# saveToCSV(df.shape[0], df.shape[1], df)
# df1 = pd.read_csv('mycsv.csv')
# print(df1)


# JSON数据
import json

"""
json.drumps（）将数据字典转化为JSON编码的字符串str
json.loads（）将字符串转化为原来的数据字典dict格式
    ensure_ascii默认值为True,无法识别汉字
    sort_keys默认值为False,不排序
    indent设置合适的缩进,美化输出效果
"""
data = {
    'Tom': {'Sex': 'M', 'Weight': 65, 'Height': 100},
    'Nina': {'Sex': 'F', 'Weight': 55, 'Height': 170},
    'Echo': {'Sex': 'F', 'Weight': 60, 'Height': 170}
}
# json_str = json.dumps(data)
# print(json_str)
# print(type(json_str))
#
# json_dict = json.loads(json_str)
# print(json_dict)
# print(type(json_dict))


# XLSX文件
"""
进行Excel文件的读写操作
xlrd、xlwt、xlsxwriter、openpyxl
openpyxl同时具有读写的功能，
"""
import os
import numpy as np
# import xlsxwriter
from openpyxl import load_workbook

if 'myXlsxFolder' not in os.listdir():  # :#如果没有 myxsxFolder 文件夹，则创建(避免重复创建#报错)
    os.mkdir('myXlsxFolder')  # 在当前目录创建文件夹
os.chdir('myXlsxFolder')  # 打开文件夹 myxlsxFolder
myFile = load_workbook('book_info.xlsx')
# print("表名:", myFile.sheetnames) # 打印工作表名称
"""
myXlsxFile = xlsxwriter.Workbook("book info.lsx")  # 创建电子表格文件 book info.xlsx
myWorkSheet = myXsxFile.add
worksheet('mySheet')  # 为电子表格 book info.xlsx 添加一个名为mySheet的工作表
myWorkSheet.write('A1'，"《三国演义》")  # 在单元格 A1写人数据 Helo World!
myXlsxFile.close()  # 关闭电子表格
"""
mySheet = myFile['豆瓣读书']  # 通过表名选择工作表#打印工作表数据行数
# print("行数:", len(list(mySheet.rows)))
# print("列数:", len(list(mySheet.columns)))  # 打印工作表数据列数


# MySQL文件的存取
import pyMySOL  # 导人PyMySQL支持包

# 使用 connect 创建数据库连接,常用参数有本地主机、用户名 root,密码PyDatabase 数据库、编码#格式
db = pyMySQL.connect(host="localhost", user="root", password="密码", db="PyDatabase", charset='utf8')
cursor = db.cursor()  # 获取游标，用它来执行数据库的操作


# try:
#     cursor.execute("Select Version()")  # 执行 SOL语句
#     data = cursor.fetchone()
#     print("Database version: %s" % data)
# finally:
#     db.close()  # 关闭数据库

# 定义打印列名的函数 print_colsname()
def print_colsname():
    cursor.execute("SHOW COLUMNS FROM Py_Data;")
    cols_name = cursor.fetchall()  # 执行 SQL语句，查询 Py_Data中的列
    # print(cols_name)
    return cols_name


# 定义打印数据的函数print alldata()
def print_alldata():
    cursor.execute("SELECT * FROM Py_Data;")  # 执行SQL语句，查询 Py Data中的所有数据
    data = cursor.fetchal1()  # 获取全部数据print(data)
    return data


# 执行SQL语句
try:
    # 删除表(防止Py_Data创建前已存在)
    cursor.execute("DROP TABLE IF EXISTS Py_Data;")  # 执行SOL删除语句 Py_Data
    # 创建表
    cursor.execute("CREATE TABLE Py_Data (username VARCHAR (10), useraddr VARCHAR (20));")
    # 插入数据
    cursor.execute("INSERT INTO Py_Data (username,useraddr) VALUES('张三','中国');")
    cursor.execute("INSERT INTO Py_Data(username,useraddr) VALUES('Lisa','美国');")
    # 打印数据
    print_alldata()
    # 删除数据
    cursor.execute("DELETE FROM Py_Data WHERE useraddr ='美国'")
    # 打印数据
    print_alldata()
    # 打印当前列名
    print_colsname()
    # 删除列
    cursor.execute("ALTER TABLE Py_Data DROP username;")
    # 添加列
    cursor.execute("ALTER TABLE Py_Data ADD COLUMN (age TINYINT(1) UNSIGNED);")  # 打印修改后的列名
    print_colsname()
    # 提交以上操作到数据库
    db.commit()
except:
    # rollback():#回滚,若出现错误则放弃执行 try,并将数据恢复到 try之前的状态
    db.rollback()
    print("Error!")
finally:
    db.close()
